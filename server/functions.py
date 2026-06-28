from __future__ import annotations

from pathlib import Path
from typing import Any

from .config import SynapseConfig
from .dedup import memory_deduplicate as _dedup
from .diff import (
    apply_update,
    detect_conflicts,
    list_pending,
    propose_update,
    reject_update,
    relink_all,
)
from .organizer import organize_vault
from .watcher import start_watcher, stop_watcher, watcher_status
from .encryption import read_text
from .index import MemoryIndex
from .memory_file import key_to_path, path_to_key, parse_memory_text, render_memory_file
from .ai_importer import (
    import_ai_export as _import_ai_export,
    ingest_text as _ingest_text,
    import_filtered_jsonl as _import_filtered_jsonl,
    import_synapse_summaries as _import_synapse_summaries,
    save_chat_memory as _save_chat_memory,
)
from .graph_builder import build_topic_graph as _build_topic_graph, deep_search as _deep_search
from .raw_archive import (
    get_raw_conversation as _get_raw,
    get_raw_chunks as _get_raw_chunks,
    search_raw_index as _search_raw_index,
)
from .merger import smart_merge_duplicates as _smart_merge
from .triage import run_triage as _run_triage
from .claude_formatter import format_claude_export as _format_claude_export
from . import background as _bg
from .scanner import scan_and_extract
from .search import memory_search


def _estimate_tokens(obj: Any) -> int:
    """Rough token estimate: 1 token ≈ 4 characters of JSON-serialised text."""
    try:
        import json as _json

        return max(1, len(_json.dumps(obj, ensure_ascii=False)) // 4)
    except Exception:
        return 0


def memory_tree(config: SynapseConfig, confirm: bool = False) -> dict[str, Any]:
    """Return a nested JSON directory tree for the configured vault."""
    vault = config.vault_path
    _ensure_vault_path(vault)
    md_files = list(vault.rglob("*.md"))
    est_tokens = len(md_files) * 80
    if not confirm:
        return {
            "warning": "memory_tree is expensive (~20k tokens for large vaults). Call with confirm=True to proceed.",
            "estimated_tokens": est_tokens,
            "file_count": len(md_files),
            "tip": "Use memory_list or memory_search instead — same info at 1/10th the cost.",
        }
    return _tree_node(vault, vault)


def memory_get(config: SynapseConfig, key: str) -> dict[str, Any]:
    """Return parsed frontmatter and content for a dot-notation memory key."""
    vault = config.vault_path
    _ensure_vault_path(vault)
    file_path = key_to_path(vault, key)
    if not file_path.exists():
        raise ValueError(f"Memory key not found: {key}")
    if not file_path.is_file() or file_path.suffix.lower() != ".md":
        raise ValueError(f"Memory key does not resolve to a Markdown file: {key}")

    frontmatter, content = parse_memory_text(read_text(config, file_path))
    result = {
        "key": frontmatter.get("key", key),
        "file_path": str(file_path.relative_to(vault)).replace("\\", "/"),
        "frontmatter": frontmatter,
        "content": content,
    }
    result["_tokens"] = _estimate_tokens(result)
    _bg.track_usage("memory_get", result["_tokens"], key=key)
    return result


def memory_search_tool(config: SynapseConfig, query: str) -> list[dict[str, Any]]:
    hits = memory_search(config, query)
    _bg.track_usage("memory_search", _estimate_tokens(hits))
    return hits


def memory_propose_update(config: SynapseConfig, patch: dict[str, Any]) -> dict[str, Any]:
    return propose_update(config, patch)


def _append_changelog(config: SynapseConfig, key: str, action: str = "updated") -> None:
    """Append a line to vault/metadata/changelog.md after every successful write."""
    from datetime import datetime
    changelog = config.vault_path / "metadata" / "changelog.md"
    changelog.parent.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    line = f"- `{ts}` — {action}: `{key}`\n"
    with changelog.open("a", encoding="utf-8") as f:
        f.write(line)


def memory_apply_update(config: SynapseConfig, patch_id: str) -> dict[str, Any]:
    result = apply_update(config, patch_id)
    if "error" not in result:
        _bg.schedule_rebuild(_rebuild_index_sync, config)
        key = result.get("key", patch_id)
        _bg.track_usage("memory_apply_update", 0, key=key, write=True)
        try:
            _append_changelog(config, key)
        except Exception:
            pass
    return result


def memory_reject_update(config: SynapseConfig, patch_id: str, reason: str = "") -> dict[str, Any]:
    return reject_update(config, patch_id, reason)


def memory_diff(config: SynapseConfig) -> list[dict[str, Any]]:
    return list_pending(config)


def memory_conflicts(config: SynapseConfig, auto_resolve: bool = False) -> list[dict[str, Any]]:
    conflicts = detect_conflicts(config)
    if not auto_resolve or not conflicts:
        return conflicts

    resolutions: list[dict[str, Any]] = []
    for conflict in conflicts:
        left_key, right_key = conflict["left"], conflict["right"]
        try:
            left = memory_get(config, left_key)
            right = memory_get(config, right_key)
        except Exception as e:
            resolutions.append({**conflict, "resolution": "error", "error": str(e)})
            continue

        # Newer file wins (ISO date string comparison is correct)
        left_date = left["frontmatter"].get("last_updated", "0000-00-00")
        right_date = right["frontmatter"].get("last_updated", "0000-00-00")
        winner_key, loser, loser_key = (
            (left_key, right, right_key) if left_date >= right_date
            else (right_key, left, left_key)
        )

        fm = dict(loser["frontmatter"])
        fm["confidence"] = "deprecated"
        new_body = f"**Superseded by `{winner_key}`** — kept for reference.\n\n{loser['content']}"
        full_text = render_memory_file(fm, new_body)
        try:
            r = propose_update(config, {"key": loser_key, "content": full_text, "merge": "replace"})
            resolutions.append({
                **conflict,
                "resolution": "proposed",
                "winner": winner_key,
                "deprecated": loser_key,
                "patch_id": r.get("patch_id"),
            })
        except Exception as e:
            resolutions.append({**conflict, "resolution": "error", "error": str(e)})

    return resolutions


def memory_scan_project(
    config: SynapseConfig,
    path: str,
    exclude_dirs: list[str] | None = None,
    background: bool = True,
) -> dict[str, Any]:
    if background:
        return _bg.start_job(
            "scan_project",
            _scan_project_sync,
            config, path, exclude_dirs,
        )
    return _scan_project_sync(config, path, exclude_dirs)


def _scan_project_sync(
    config: SynapseConfig,
    path: str,
    exclude_dirs: list[str] | None = None,
) -> dict[str, Any]:
    result = scan_and_extract(config, path, exclude_dirs=exclude_dirs)
    if "error" in result and "proposals" not in result:
        return result

    patch_ids: list[Any] = []
    errors: list[dict[str, str]] = []
    for proposal in result.get("proposals", []):
        try:
            r = propose_update(config, proposal)
            patch_ids.append(r["patch_id"])
        except Exception as exc:
            errors.append({"key": proposal.get("key", "?"), "error": str(exc)})

    return {
        "project_name": result.get("project_name", ""),
        "detected": result.get("detected", {}),
        "files_analyzed": result.get("file_count", 0),
        "function_nodes": result.get("function_nodes", 0),
        "stale_removed": result.get("stale_removed", 0),
        "patches_proposed": len(patch_ids),
        "patch_ids": patch_ids,
        "errors": errors,
        "next_step": "Call memory_diff to review patches, then memory_apply_update per patch_id.",
    }


def memory_ingest_text(
    config: SynapseConfig, text: str, label: str = "[pasted text]"
) -> dict[str, Any]:
    result = _ingest_text(config, text, label)
    if "error" in result:
        return result

    patch_ids: list[Any] = []
    errors: list[dict[str, str]] = []
    for proposal in result.get("proposals", []):
        try:
            r = propose_update(config, proposal)
            patch_ids.append(r["patch_id"])
        except Exception as exc:
            errors.append({"key": proposal.get("key", "?"), "error": str(exc)})

    return {
        "patches_proposed": len(patch_ids),
        "patch_ids": patch_ids,
        "errors": errors,
        "next_step": "Call memory.diff to review, then memory.apply_update per patch_id.",
    }


def memory_smart_merge(
    config: SynapseConfig, dry_run: bool = True, threshold: float = 0.93
) -> dict[str, Any]:
    return _smart_merge(config, dry_run=dry_run, threshold=threshold)


def memory_import_ai_export(
    config: SynapseConfig,
    path: str,
    owner_name: str | None = None,
    resume_failed: bool = False,
) -> dict[str, Any]:
    result = _import_ai_export(config, path, owner_name, resume_failed=resume_failed)
    if "error" in result and "proposals" not in result:
        return result

    patch_ids: list[Any] = []
    errors: list[dict[str, str]] = []
    for proposal in result.get("proposals", []):
        try:
            r = propose_update(config, proposal)
            patch_ids.append(r["patch_id"])
        except Exception as exc:
            errors.append({"key": proposal.get("key", "?"), "error": str(exc)})

    return {
        "provider": result.get("provider", "unknown"),
        "chunks_processed": result.get("chunks_processed", 0),
        "owner_detected": result.get("owner_detected"),
        "identity_warning": result.get("identity_warning"),
        "failed_chunks": result.get("failed_chunks", 0),
        "resume_file": result.get("resume_file"),
        "patches_proposed": len(patch_ids),
        "patch_ids": patch_ids,
        "errors": errors,
        "next_step": "Call memory.diff to review, then memory.apply_update per patch_id.",
        "triage_warning": (
            "Quality filtering was bypassed. For Claude exports, use the proper pipeline instead: "
            "memory_format_claude_export → memory_triage → memory_import_filtered_jsonl. "
            "This ensures sensitive content (secrets, financial, health data) is redacted before import."
        ),
    }


def memory_import_filtered_jsonl(
    config: SynapseConfig,
    filtered_jsonl_folder: str,
    blacklist_file: str | None = None,
    redflag_file: str | None = None,
    owner_name: str | None = None,
) -> dict[str, Any]:
    result = _import_filtered_jsonl(
        config, filtered_jsonl_folder, blacklist_file, redflag_file, owner_name
    )
    if "error" in result and "proposals" not in result:
        return result

    patch_ids: list[Any] = []
    errors: list[dict[str, str]] = []
    for proposal in result.get("proposals", []):
        try:
            r = propose_update(config, proposal)
            patch_ids.append(r["patch_id"])
        except Exception as exc:
            errors.append({"key": proposal.get("key", "?"), "error": str(exc)})

    _bg.schedule_rebuild(_rebuild_index_sync, config)
    return {
        "provider": result.get("provider", "gemma"),
        "model": result.get("model"),
        "chunks_processed": result.get("chunks_processed", 0),
        "failed_chunks": result.get("failed_chunks", 0),
        "owner_detected": result.get("owner_detected"),
        "patches_proposed": len(patch_ids),
        "patch_ids": patch_ids,
        "errors": errors,
        "next_step": "Call memory.diff to review, then memory.apply_update per patch_id.",
    }


def memory_import_synapse_summaries(
    config: SynapseConfig,
    summaries_folder: str,
    owner_name: str | None = None,
) -> dict[str, Any]:
    result = _import_synapse_summaries(config, summaries_folder, owner_name)
    _bg.schedule_rebuild(_rebuild_index_sync, config)
    return result


def memory_organize_vault(config: SynapseConfig) -> dict[str, Any]:
    result = organize_vault(config)
    _bg.schedule_rebuild(_rebuild_index_sync, config)
    return result


def memory_relink_all(config: SynapseConfig) -> dict[str, Any]:
    result = relink_all(config)
    _bg.schedule_rebuild(_rebuild_index_sync, config)
    return result


def memory_start_watcher(config: SynapseConfig, path: str) -> dict[str, Any]:
    return start_watcher(config, path)


def memory_stop_watcher() -> dict[str, Any]:
    return stop_watcher()


def memory_watcher_status() -> dict[str, Any]:
    return watcher_status()


def memory_dedup(config: SynapseConfig, auto_clean: bool = False) -> dict[str, Any]:
    return _dedup(config, auto_clean=auto_clean)


def memory_list_folder(config: SynapseConfig, folder: str = "") -> dict[str, Any]:
    """Shallow listing of one vault folder — cheap alternative to memory_tree."""
    vault = config.vault_path
    _ensure_vault_path(vault)
    target = vault / folder if folder else vault
    if not target.exists() or not target.is_dir():
        raise ValueError(f"Folder not found: {folder!r}")
    files, folders = [], []
    for item in sorted(target.iterdir(), key=lambda p: (p.is_file(), p.name.lower())):
        if item.name.startswith("_") or item.name.startswith("."):
            continue
        if item.is_dir():
            folders.append(item.name)
        elif item.suffix.lower() == ".md":
            files.append(path_to_key(vault, item))
    return {"folder": folder or ".", "subfolders": folders, "keys": files}


def memory_context(config: SynapseConfig) -> dict[str, Any]:
    """Return core identity context + vault health check in one call for conversation start."""
    result: dict[str, Any] = {}
    for key in ("identity.profile", "identity.communication", "identity.location"):
        try:
            result[key] = memory_get(config, key)
        except Exception:
            result[key] = None

    # Lightweight folder index — always present even on empty vault
    for folder in ("life", "work", "patterns", "projects"):
        try:
            result[f"_index.{folder}"] = memory_list_folder(config, folder)["keys"]
        except Exception:
            result[f"_index.{folder}"] = []

    # Run dedup automatically — Jaccard only, no API cost
    try:
        dedup = _dedup(config, auto_clean=False)
        health: dict[str, Any] = {
            "total_files": sum(
                1 for _ in config.vault_path.rglob("*.md") if not _.name.startswith("_")
            )
        }
        issues = []
        if dedup.get("stray_files"):
            issues.append(f"{len(dedup['stray_files'])} stray files")
        if dedup.get("thin_files"):
            issues.append(f"{len(dedup['thin_files'])} thin stubs")
        if dedup.get("duplicate_groups"):
            issues.append(f"{len(dedup['duplicate_groups'])} duplicate groups")
        health["issues"] = issues
        health["clean"] = len(issues) == 0
        result["_vault_health"] = health
    except Exception:
        pass

    result["_write_mode"] = config.write_mode
    result["_tokens"] = _estimate_tokens(result)
    _bg.track_usage("memory_context", result["_tokens"])
    return result


def memory_get_raw(config: SynapseConfig, chat_id: str) -> dict[str, Any]:
    return _get_raw(config, chat_id)


def memory_get_raw_chunks(
    config: SynapseConfig, chat_id: str, query: str, top_k: int = 3, window: int = 8
) -> dict[str, Any]:
    result = _get_raw_chunks(config, chat_id, query, top_k=top_k, window=window)
    if isinstance(result, dict):
        result["_tokens"] = _estimate_tokens(result)
    return result


def memory_search_raw(config: SynapseConfig, query: str, top_k: int = 10) -> list[dict[str, Any]]:
    return _search_raw_index(config, query, top_k=top_k)


def memory_code_search(
    config: SynapseConfig,
    query: str,
    project: str = "",
    limit: int = 8,
) -> list[dict[str, Any]]:
    """Hybrid FTS5 + semantic search over code nodes indexed by memory_scan_project."""
    from .code_index import search_code, list_projects

    if not config.vault_path or not (config.vault_path / "_code_index.db").exists():
        return [{"error": "No code index found. Run memory_scan_project first."}]
    results = search_code(
        config.vault_path,
        config.gemini_api_key or None,
        query,
        project=project,
        limit=limit,
    )
    if not results:
        projects = list_projects(config.vault_path)
        return [{"info": "No results.", "indexed_projects": projects}]
    return results


def memory_code_stats(config: SynapseConfig, project: str = "") -> dict[str, Any]:
    """Stats for indexed code projects. Pass project slug to drill in."""
    from .code_index import list_projects, project_stats

    if not config.vault_path or not (config.vault_path / "_code_index.db").exists():
        return {"error": "No code index found. Run memory_scan_project first."}
    projects = list_projects(config.vault_path)
    if project:
        if project not in projects:
            return {"error": f"Project {project!r} not indexed.", "indexed_projects": projects}
        return {"project": project, **project_stats(config.vault_path, project)}
    return {"indexed_projects": projects}


def memory_save_chat(
    config: SynapseConfig,
    title: str,
    summary: str,
    key_facts: list[str],
    decisions: list[str],
    tags: list[str],
    keywords: str = "",
    categories: list[str] | None = None,
    chat_id: str | None = None,
) -> dict[str, Any]:
    result = _save_chat_memory(
        config,
        title,
        summary,
        key_facts,
        decisions,
        tags,
        keywords=keywords,
        categories=categories,
        chat_id=chat_id,
    )
    if "error" not in result:
        _bg.schedule_rebuild(_rebuild_index_sync, config)
    return result


_CHAT_SECTIONS = [
    "Deep Summary",
    "Key Facts",
    "Decisions Made",
    "Problems Solved",
    "Technical Details",
    "Files Ingested",
    "References",
    "Next Steps",
    "Timeline",
]


def _build_chat_template(title: str, chat_id: str, tags: list[str]) -> str:
    import datetime as _dt
    today = _dt.date.today().isoformat()
    now = _dt.datetime.now().strftime("%H:%M")
    tags_yaml = "\n".join(f"- {t}" for t in tags) if tags else "- general"
    cats_yaml = "- general"
    frontmatter = (
        f"---\n"
        f"key: chats.{chat_id}\n"
        f"type: chat_summary\n"
        f"title: {title}\n"
        f"status: draft\n"
        f"created: {today}\n"
        f"source: claude_session\n"
        f"categories:\n{cats_yaml}\n"
        f"tags:\n{tags_yaml}\n"
        f"related: []\n"
        f"---\n"
    )
    body = f"# {title}\n\n"
    body += f"**Status:** Draft  \n**Started:** {today} {now}\n\n"
    for section in _CHAT_SECTIONS:
        body += f"## {section}\n\n_(empty)_\n\n"
    return frontmatter + body


def _patch_chat_section(text: str, section: str, new_items: list[str] | str) -> str:
    """Append items to a named ## section in the chat markdown. Creates the section if missing."""
    heading = f"## {section}"
    lines = text.splitlines(keepends=True)
    section_start = None
    section_end = None
    for i, line in enumerate(lines):
        if line.strip() == heading:
            section_start = i
        elif section_start is not None and line.startswith("## ") and i > section_start:
            section_end = i
            break
    if section_start is None:
        text = text.rstrip() + f"\n\n{heading}\n\n"
        lines = text.splitlines(keepends=True)
        section_start = len(lines) - 1
        section_end = None

    # Build new content block
    if isinstance(new_items, str):
        addition = new_items.strip() + "\n"
    else:
        addition = "\n".join(f"- {item}" for item in new_items if item.strip()) + "\n"

    # Find insertion point: just before next ## or end of file
    insert_at = section_end if section_end is not None else len(lines)

    # Remove the _(empty)_ placeholder if present
    block = lines[section_start + 1: insert_at]
    cleaned = [l for l in block if l.strip() != "_(empty)_"]
    if cleaned and not cleaned[-1].strip():
        cleaned = cleaned[:-1]  # drop trailing blank

    new_block = cleaned + [addition, "\n"]
    lines[section_start + 1: insert_at] = new_block
    return "".join(lines)


def memory_start_chat(
    config: SynapseConfig,
    title: str,
    initial_topic: str = "",
) -> dict[str, Any]:
    """
    Call at the START of every conversation to create a detailed living chat record.
    Returns a chat_id — pass it to memory_update_chat throughout the session, and
    to memory_finalize_chat at the end.

    The record is saved immediately as a draft in vault/chats/ and gets updated in-place
    as the conversation progresses.
    """
    import uuid as _uuid
    chat_id = str(_uuid.uuid4())
    vault = config.vault_path
    chats_dir = vault / "chats"
    chats_dir.mkdir(parents=True, exist_ok=True)

    template = _build_chat_template(title, chat_id, tags=["general"])
    chat_path = chats_dir / f"{chat_id}.md"
    chat_path.write_text(template, encoding="utf-8")

    if initial_topic:
        text = chat_path.read_text(encoding="utf-8")
        text = _patch_chat_section(text, "Deep Summary", initial_topic)
        chat_path.write_text(text, encoding="utf-8")

    # Update FTS index
    try:
        from .index import MemoryIndex
        idx = MemoryIndex(vault, lambda p: read_text(config, p))
        idx.upsert_file(chat_path)
    except Exception:
        pass

    _bg.track_usage("memory_start_chat", 0, key=f"chats.{chat_id}", write=True)
    return {
        "chat_id": chat_id,
        "key": f"chats.{chat_id}",
        "file": f"chats/{chat_id}.md",
        "status": "draft_created",
        "instructions": (
            "Record created. Call memory_update_chat(chat_id, ...) whenever something notable happens. "
            "Call memory_finalize_chat(chat_id, summary) at conversation end."
        ),
    }


def memory_update_chat(
    config: SynapseConfig,
    chat_id: str,
    key_facts: list[str] | None = None,
    decisions: list[str] | None = None,
    problems_solved: list[str] | None = None,
    technical_details: str = "",
    references: list[str] | None = None,
    next_steps: list[str] | None = None,
    timeline: list[str] | None = None,
    deep_summary: str = "",
    tags: list[str] | None = None,
) -> dict[str, Any]:
    """
    Update the living chat record mid-conversation. Call whenever something notable happens:
    a decision is made, a fact is established, a problem is solved, code is written, etc.
    Pass only the sections that have new content — everything else stays unchanged.

    All list fields are APPENDED to the existing section (not replaced).
    deep_summary and technical_details replace their section if provided.
    """
    vault = config.vault_path
    chat_path = vault / "chats" / f"{chat_id}.md"
    if not chat_path.exists():
        return {"error": f"Chat {chat_id} not found. Call memory_start_chat first."}

    text = chat_path.read_text(encoding="utf-8")

    updates = {
        "Key Facts": key_facts,
        "Decisions Made": decisions,
        "Problems Solved": problems_solved,
        "References": references,
        "Next Steps": next_steps,
        "Timeline": timeline,
    }
    sections_updated = []
    for section, items in updates.items():
        if items:
            text = _patch_chat_section(text, section, items)
            sections_updated.append(section)

    if technical_details.strip():
        text = _patch_chat_section(text, "Technical Details", technical_details.strip())
        sections_updated.append("Technical Details")

    if deep_summary.strip():
        text = _patch_chat_section(text, "Deep Summary", deep_summary.strip())
        sections_updated.append("Deep Summary")

    # Update tags in frontmatter if provided
    if tags:
        import re
        tag_block = "tags:\n" + "\n".join(f"- {t}" for t in tags)
        text = re.sub(r"tags:\n(- .+\n)*", tag_block + "\n", text, count=1)

    chat_path.write_text(text, encoding="utf-8")

    try:
        from .index import MemoryIndex
        idx = MemoryIndex(vault, lambda p: read_text(config, p))
        idx.upsert_file(chat_path)
    except Exception:
        pass

    return {
        "chat_id": chat_id,
        "sections_updated": sections_updated,
        "status": "updated",
    }


def memory_finalize_chat(
    config: SynapseConfig,
    chat_id: str,
    summary: str,
    tags: list[str] | None = None,
) -> dict[str, Any]:
    """
    Call at the END of every conversation to mark the chat record complete.
    Writes the final deep summary, links any ingested files, updates tags/categories,
    marks status as 'complete', and triggers index rebuild.

    - chat_id: returned by memory_start_chat
    - summary: detailed 3-8 sentence summary of the full conversation
    - tags: topic tags (auto-derived from session activity if omitted)
    """
    import re

    vault = config.vault_path
    chat_path = vault / "chats" / f"{chat_id}.md"
    if not chat_path.exists():
        return {"error": f"Chat {chat_id} not found."}

    chat_key = f"chats.{chat_id}"
    text = chat_path.read_text(encoding="utf-8")

    # Write final summary
    text = _patch_chat_section(text, "Deep Summary", summary.strip())

    # Auto-derive tags from session if not provided
    stats = _bg.get_session_stats()
    if not tags:
        touched = stats.get("keys_read", []) + stats.get("keys_written", [])
        folders = list(dict.fromkeys(
            k.split(".")[0] for k in touched
            if "." in k and not k.startswith("files.") and not k.startswith("chats.")
        ))
        tags = folders or ["general"]

    # Update frontmatter: status → complete, tags
    text = re.sub(r"status: draft", "status: complete", text, count=1)
    tag_block = "tags:\n" + "\n".join(f"- {t}" for t in tags)
    text = re.sub(r"tags:\n(- .+\n)*", tag_block + "\n", text, count=1)

    # Append files ingested section
    file_keys = _bg.get_unlinked_file_keys()
    if file_keys:
        file_list = [f"[[{k}]]" for k in file_keys]
        text = _patch_chat_section(text, "Files Ingested", file_list)

    chat_path.write_text(text, encoding="utf-8")

    # Link each file back to this chat via its related frontmatter field
    for file_key in file_keys:
        try:
            file_mem = memory_get(config, file_key)
            if "error" in file_mem:
                continue
            fm = dict(file_mem.get("frontmatter", {}))
            related = list(fm.get("related", []))
            if chat_key not in related:
                related.append(chat_key)
            fm["related"] = related
            propose_update(config, {
                "key": file_key,
                "content": file_mem.get("content", ""),
                "merge": "replace",
                "frontmatter": fm,
                "reason": f"Linked to chat: {chat_id}",
            })
        except Exception:
            pass

    if file_keys:
        _bg.mark_files_linked(file_keys)

    # Strip leftover _(empty)_ placeholders from sections never touched
    import re as _re
    text = chat_path.read_text(encoding="utf-8")
    text = _re.sub(r"\n_\(empty\)_\n", "\n", text)
    chat_path.write_text(text, encoding="utf-8")

    # Update FTS index
    try:
        from .index import MemoryIndex
        idx = MemoryIndex(vault, lambda p: read_text(config, p))
        idx.upsert_file(chat_path)
    except Exception:
        pass

    _bg.schedule_rebuild(_rebuild_index_sync, config)
    _bg.track_usage("memory_finalize_chat", 0, key=chat_key, write=True)

    return {
        "chat_id": chat_id,
        "key": chat_key,
        "file": f"chats/{chat_id}.md",
        "status": "complete",
        "files_linked": file_keys,
        "tags": tags,
        "note": (
            f"Chat finalized. "
            + (f"Linked {len(file_keys)} file(s)." if file_keys else "")
        ).strip(),
    }


def memory_quick_save_chat(
    config: SynapseConfig,
    title: str,
    summary: str,
    key_facts: list[str] | None = None,
    decisions: list[str] | None = None,
) -> dict[str, Any]:
    """
    One-shot fallback: create + immediately finalize a chat record in one call.
    Use memory_start_chat / memory_update_chat / memory_finalize_chat for richer records.
    """
    r = memory_start_chat(config, title=title)
    if "error" in r:
        return r
    chat_id = r["chat_id"]
    if key_facts or decisions:
        memory_update_chat(config, chat_id, key_facts=key_facts, decisions=decisions)
    return memory_finalize_chat(config, chat_id, summary=summary)


def memory_build_graph(config: SynapseConfig, top_k: int = 8, background: bool = False) -> dict[str, Any]:
    if background:
        return _bg.start_job("build_graph", _build_graph_with_progress, config, top_k)
    return _build_topic_graph(config, top_k=top_k)


def _build_graph_with_progress(config: SynapseConfig, top_k: int) -> dict[str, Any]:
    import threading as _threading

    _cancel = _bg.get_stop_flag("build_graph")
    if _cancel.is_set():
        return {"status": "cancelled"}

    _bg.update_progress("build_graph", 5, "loading chat nodes…")

    _stop = _threading.Event()
    def _heartbeat():
        import time as _time
        milestones = [(5, 20, "building edges…"), (15, 50, "computing links…"), (30, 70, "writing graph…")]
        start = _time.time()
        while not _stop.is_set():
            elapsed = _time.time() - start
            for secs, pct, label in milestones:
                if elapsed >= secs:
                    _bg.update_progress("build_graph", pct, label)
            _stop.wait(3)

    t = _threading.Thread(target=_heartbeat, daemon=True)
    t.start()
    try:
        result = _build_topic_graph(config, top_k=top_k)
    finally:
        _stop.set()

    _bg.update_progress("build_graph", 95, "writing graph file…")
    return result


def memory_deep_search(
    config: SynapseConfig, query: str, depth: int = 2, top_k: int = 8
) -> list[dict[str, Any]]:
    return _deep_search(config, query, depth=depth, top_k=top_k)


def memory_auto(config: SynapseConfig, task: str) -> dict[str, Any]:
    """
    Smart retrieval dispatcher. Always loads context, searches active vault,
    and escalates to deep search only when no vault result clears the confidence
    threshold (score >= 0.7). One strong hit stops the chain; weak hits escalate.
    """
    from .search import CLAUDE_ANALYSIS_THRESHOLD

    result: dict[str, Any] = {}

    # Tier 1: always
    result["context"] = memory_context(config)

    # Tier 2: active vault search
    vault_hits = memory_search_tool(config, task)
    result["vault_results"] = vault_hits

    # Tier 3: escalate only when no hit is confident enough
    best_score = max((r.get("score", 0.0) for r in vault_hits), default=0.0)
    if best_score < CLAUDE_ANALYSIS_THRESHOLD:
        result["deep_results"] = _deep_search(config, task)

    result["_tokens"] = _estimate_tokens(result)
    _bg.track_usage("memory_auto", result["_tokens"])
    return result


def memory_commit(config: SynapseConfig, patch: dict[str, Any]) -> dict[str, Any]:
    """
    Write a memory patch using the configured write_mode.
    - review (default): proposes a diff for human approval, same as memory_propose_update.
    - auto: proposes then immediately applies — no confirmation needed.
    """
    proposed = memory_propose_update(config, patch)
    if config.write_mode != "auto":
        return proposed
    patch_id = proposed.get("patch_id")
    if not patch_id:
        return proposed
    return memory_apply_update(config, patch_id)


def memory_format_claude_export(
    config: SynapseConfig,
    export_folder: str,
    output_folder: str = "",
    write_markdown: bool = True,
) -> dict[str, Any]:
    """Format a Claude.ai export into monthly JSONL files ready for memory_triage."""
    out = output_folder or str(config.root_path / "synapse_extracted")
    return _format_claude_export(
        export_folder=export_folder,
        output_folder=out,
        write_markdown=write_markdown,
    )


def memory_triage(
    config: SynapseConfig,
    input_folder: str,
    output_folder: str = "",
    force_review: bool = False,
    openrouter_model: str = "",
    groq_model: str = "",
    workers: int = 3,
) -> dict[str, Any]:
    """Run the AI triage pipeline on a conversations_jsonl folder.

    Preferred: OPENROUTER_API_KEY + GROQ_API_KEY (sensitive content stays off Google).
    Fallback: when both are absent, uses GEMINI_API_KEY with a privacy warning.
    """
    import os

    or_key = os.environ.get("OPENROUTER_API_KEY", "")
    groq_key = os.environ.get("GROQ_API_KEY", "")
    gemini_key = config.gemini_api_key or ""

    if not or_key and not groq_key and not gemini_key:
        return {"error": "No API keys set. Need OPENROUTER_API_KEY + GROQ_API_KEY (preferred) or GEMINI_API_KEY (fallback)."}

    out = output_folder or str(config.root_path / "synapse_filtered_chats")

    kwargs: dict[str, Any] = {}
    if openrouter_model:
        kwargs["openrouter_model"] = openrouter_model
    if groq_model:
        kwargs["groq_model"] = groq_model

    return _run_triage(
        input_folder=input_folder,
        output_folder=out,
        openrouter_api_key=or_key,
        groq_api_key=groq_key,
        gemini_api_key=gemini_key,
        force_review=force_review,
        workers=workers,
        **kwargs,
    )


def rebuild_index(config: SynapseConfig, background: bool = False) -> dict[str, Any]:
    if background:
        return _bg.start_job("rebuild_index", _rebuild_index_sync, config)
    return _rebuild_index_sync(config)


def _rebuild_index_sync(config: SynapseConfig) -> dict[str, str]:
    import threading as _threading

    _cancel = _bg.get_stop_flag("rebuild_index")
    if _cancel.is_set():
        return {"status": "cancelled"}

    _bg.update_progress("rebuild_index", 5, "starting rebuild…")

    # Time-based progress heartbeat: MemoryIndex.rebuild() is opaque so we estimate
    # based on elapsed time (30s→20%, 60s→40%, 120s→60%, 240s→75%)
    _stop = _threading.Event()
    def _heartbeat():
        import time as _time
        milestones = [(10, 10, "loading vault files…"), (30, 20, "indexing content…"),
                      (60, 40, "computing embeddings…"), (120, 60, "embedding large files…"),
                      (240, 75, "almost done…")]
        start = _time.time()
        while not _stop.is_set():
            elapsed = _time.time() - start
            for secs, pct, label in milestones:
                if elapsed >= secs:
                    _bg.update_progress("rebuild_index", pct, label)
            _stop.wait(5)

    t = _threading.Thread(target=_heartbeat, daemon=True)
    t.start()
    try:
        MemoryIndex(config.vault_path, lambda path: read_text(config, path)).rebuild(
            api_key=config.gemini_api_key or None
        )
    finally:
        _stop.set()

    _bg.update_progress("rebuild_index", 95, "flushing index…")
    return {"status": "rebuilt", "db_path": str((config.vault_path / "_index.db").resolve())}


def memory_index_status() -> dict[str, Any]:
    """Return status of all background jobs (rebuild_index, build_graph, full_import)."""
    return _bg.get_status()


def memory_stop_job(job_name: str = "") -> dict[str, Any]:
    """
    Cancel a running background job. Pass job_name to stop one job
    (rebuild_index, build_graph, full_import), or leave empty to cancel all.
    Also cancels any pending auto-rebuild timer.
    """
    return _bg.cancel_job(job_name)


def memory_full_import(
    config: SynapseConfig,
    export_folder: str,
    owner_name: str = "",
    skip_triage: bool = False,
    background_rebuild: bool = True,
) -> dict[str, Any]:
    """
    One-command pipeline: format → triage → import_filtered → rebuild_index → build_graph.

    Steps:
      1. memory_format_claude_export  — converts conversations.json → monthly JSONL
      2. memory_triage                — AI filter: keep / skip / redflag
      3. memory_import_filtered_jsonl — imports kept chats into vault (Gemini)
      4. rebuild_index                — refreshes SQLite FTS5 index (background by default)
      5. memory_build_graph           — wires new chats into topic graph (background)

    skip_triage=True jumps straight from format → import_filtered (no privacy filter).
    background_rebuild=True (default) runs steps 4 & 5 in background threads so this
    call returns immediately after the import. Poll memory_index_status for progress.
    """
    JOB = "full_import"
    _cancel = _bg.get_stop_flag(JOB)
    _cancel.clear()
    _bg.update_progress(JOB, 0, "starting pipeline…")
    progress: list[str] = []

    # Step 1: format (0 → 20%)
    fmt_out = str(config.root_path / "synapse_extracted")
    _bg.update_progress(JOB, 2, "step 1/5: formatting export…")
    progress.append("Step 1: formatting Claude export…")
    fmt_result = _format_claude_export(
        export_folder=export_folder,
        output_folder=fmt_out,
        write_markdown=True,
    )
    if "error" in fmt_result:
        return {"error": fmt_result["error"], "step_failed": "format", "progress": progress}
    jsonl_folder = fmt_result["jsonl_folder"]
    _bg.update_progress(JOB, 20, f"step 1/5 done: {fmt_result['conversations_extracted']} conversations")
    progress.append(f"Step 1 done: {fmt_result['conversations_extracted']} conversations extracted")

    if _cancel.is_set():
        return {"status": "cancelled", "step_cancelled_before": "triage", "progress": progress}

    # Step 2: triage (20 → 50%)
    filtered_jsonl_folder = jsonl_folder
    triage_result: dict[str, Any] | None = None
    if not skip_triage:
        import os
        or_key = os.environ.get("OPENROUTER_API_KEY", "")
        groq_key = os.environ.get("GROQ_API_KEY", "")
        gemini_key = config.gemini_api_key or ""
        if not or_key and not groq_key and not gemini_key:
            return {
                "error": "Triage requires OPENROUTER_API_KEY + GROQ_API_KEY or GEMINI_API_KEY. Set skip_triage=True to bypass.",
                "step_failed": "triage",
                "progress": progress,
            }
        _bg.update_progress(JOB, 22, "step 2/5: triaging (keep/skip/redflag)…")
        progress.append("Step 2: triaging (keep/skip/redflag)…")
        triage_out = str(config.root_path / "synapse_filtered_chats")
        triage_result = _run_triage(
            input_folder=jsonl_folder,
            output_folder=triage_out,
            openrouter_api_key=or_key,
            groq_api_key=groq_key,
            gemini_api_key=gemini_key,
        )
        if "error" in triage_result:
            return {"error": triage_result["error"], "step_failed": "triage", "progress": progress}
        filtered_jsonl_folder = triage_result["filtered_jsonl_folder"]
        _bg.update_progress(JOB, 50, f"step 2/5 done: kept={triage_result['kept']}, redflags={triage_result['redflagged']}")
        progress.append(f"Step 2 done: kept={triage_result['kept']}, skipped={triage_result['skipped']}, redflagged={triage_result['redflagged']}")
    else:
        _bg.update_progress(JOB, 50, "step 2/5: skipped (skip_triage=True)")
        progress.append("Step 2: skipped (skip_triage=True)")

    if _cancel.is_set():
        return {"status": "cancelled", "step_cancelled_before": "import", "progress": progress}

    # Step 3: import (50 → 75%)
    _bg.update_progress(JOB, 52, "step 3/5: importing kept chats into vault…")
    progress.append("Step 3: importing kept chats into vault…")
    import_result = _import_filtered_jsonl(
        config, filtered_jsonl_folder, owner_name=owner_name or None
    )
    if "error" in import_result and "proposals" not in import_result:
        return {"error": import_result["error"], "step_failed": "import", "progress": progress}

    patch_ids: list[Any] = []
    for proposal in import_result.get("proposals", []):
        try:
            r = propose_update(config, proposal)
            patch_ids.append(r["patch_id"])
        except Exception:
            pass
    _bg.update_progress(JOB, 75, f"step 3/5 done: {len(patch_ids)} patches proposed")
    progress.append(f"Step 3 done: {len(patch_ids)} patches proposed")

    if _cancel.is_set():
        return {"status": "cancelled", "step_cancelled_before": "rebuild", "progress": progress}

    # Steps 4 & 5: rebuild index + build graph (75 → 90%)
    _bg.update_progress(JOB, 78, "step 4/5: starting rebuild_index…")
    rebuild_result = rebuild_index(config, background=background_rebuild)
    _bg.update_progress(JOB, 85, "step 5/5: starting build_graph…")
    graph_result = memory_build_graph(config, background=background_rebuild)
    _bg.update_progress(JOB, 90, "steps 4+5 dispatched — waiting for index + graph")
    if background_rebuild:
        progress.append("Steps 4+5: rebuild_index + build_graph started in background — poll memory_index_status")
    else:
        _bg.update_progress(JOB, 100, "complete")
        progress.append("Steps 4+5: rebuild_index + build_graph complete")

    return {
        "status": "pipeline_complete",
        "progress": progress,
        "format": fmt_result,
        "triage": triage_result,
        "patches_proposed": len(patch_ids),
        "patch_ids": patch_ids,
        "rebuild_index": rebuild_result,
        "build_graph": graph_result,
        "next_step": (
            "Call memory_index_status to check background jobs, then memory_diff to review patches."
            if background_rebuild
            else "Call memory_diff to review patches, then memory_apply_update per patch_id."
        ),
    }


def memory_vault_diff(config: SynapseConfig, since: str = "", limit: int = 50) -> dict[str, Any]:
    """List vault files modified after a date (ISO format: '2025-06-01'). Leave since empty to list all, newest first."""
    from datetime import datetime, timezone

    vault = config.vault_path
    _ensure_vault_path(vault)

    cutoff_ts: float | None = None
    if since:
        try:
            dt = datetime.fromisoformat(since)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            cutoff_ts = dt.timestamp()
        except ValueError:
            return {"error": f"Invalid date: {since!r} — use ISO format like '2025-06-01' or '2025-06-01T12:00:00'."}

    changed: list[dict[str, Any]] = []
    for path in sorted(vault.rglob("*.md"), key=lambda p: p.stat().st_mtime, reverse=True):
        if path.name.startswith("_"):
            continue
        mtime = path.stat().st_mtime
        if cutoff_ts and mtime < cutoff_ts:
            continue
        rel = path.relative_to(vault)
        changed.append({
            "key": path_to_key(vault, path),
            "folder": rel.parts[0] if len(rel.parts) > 1 else "root",
            "modified": datetime.fromtimestamp(mtime, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        })
        if len(changed) >= limit:
            break

    return {"since": since or "all time", "files_changed": len(changed), "changes": changed}


def memory_apply_all(
    config: SynapseConfig, folder: str = "", dry_run: bool = False
) -> dict[str, Any]:
    """Apply all pending patches at once. Filter by vault folder (e.g. 'life'). dry_run=True previews without writing."""
    pending = list_pending(config)
    if folder:
        pending = [p for p in pending if str(p.get("key", "")).startswith(f"{folder}.")]

    if dry_run:
        return {
            "dry_run": True,
            "would_apply": len(pending),
            "patches": [{"patch_id": p["patch_id"], "key": p.get("key")} for p in pending],
        }

    applied: list[str] = []
    errors: list[dict[str, str]] = []
    for patch in pending:
        try:
            apply_update(config, patch["patch_id"])
            key = patch.get("key", patch["patch_id"])
            applied.append(patch["patch_id"])
            try:
                _append_changelog(config, key)
            except Exception:
                pass
        except Exception as e:
            errors.append({"patch_id": patch["patch_id"], "error": str(e)})

    if applied:
        _bg.schedule_rebuild(_rebuild_index_sync, config)
        _bg.track_usage("memory_apply_all", 0, write=True)

    return {"applied": len(applied), "errors": errors, "patch_ids": applied}


def memory_fix_frontmatter(
    config: SynapseConfig, dry_run: bool = True
) -> dict[str, Any]:
    """
    Find vault files with missing required frontmatter fields and propose patches to fix them.
    dry_run=True (default) reports problems without writing. dry_run=False proposes patches.
    """
    from .memory_file import new_frontmatter

    REQUIRED = {"type", "weight", "confidence", "scope"}
    vault = config.vault_path
    _ensure_vault_path(vault)

    problems: list[dict[str, Any]] = []
    patch_ids: list[str] = []

    for path in sorted(vault.rglob("*.md")):
        if path.name.startswith("_"):
            continue
        try:
            text = read_text(config, path)
            fm, body = parse_memory_text(text)
            missing = REQUIRED - set(fm.keys())
            if not missing:
                continue
            key = str(fm.get("key") or path_to_key(vault, path))
            problems.append({"key": key, "missing_fields": sorted(missing)})

            if not dry_run:
                # Build full frontmatter from defaults, then overlay existing valid values
                defaults = new_frontmatter(key=key)
                merged_fm: dict[str, Any] = dict(defaults)
                for field, value in fm.items():
                    if value is not None and value != "":
                        merged_fm[field] = value
                try:
                    # Pass frontmatter dict + body separately so _build_after_text
                    # merges them correctly (passing full rendered text as content
                    # would make it appear as body, not frontmatter)
                    r = propose_update(config, {
                        "key": key,
                        "content": body,
                        "merge": "replace",
                        "frontmatter": merged_fm,
                    })
                    patch_ids.append(r["patch_id"])
                except Exception:
                    pass
        except Exception:
            continue

    return {
        "dry_run": dry_run,
        "files_with_issues": len(problems),
        "problems": problems,
        "patches_proposed": len(patch_ids),
        "patch_ids": patch_ids,
        "next_step": "Call memory_apply_all() to apply all fixes at once." if patch_ids else "",
    }


def memory_multi_search(
    config: SynapseConfig, queries: list[str], top_k: int = 4
) -> dict[str, Any]:
    """
    Fan-out search: run multiple queries in parallel and merge results, deduplicating by key.
    Returns a ranked list where results appearing in multiple queries score higher.
    """
    import threading as _threading

    results_by_query: dict[str, list[dict[str, Any]]] = {}
    lock = _threading.Lock()

    def _run(q: str) -> None:
        hits = memory_search(config, q)[:top_k]
        with lock:
            results_by_query[q] = hits

    threads = [_threading.Thread(target=_run, args=(q,), daemon=True) for q in queries]
    for t in threads:
        t.start()
    for t in threads:
        t.join(timeout=10)

    # Merge: accumulate score per key, keep best snippet
    merged: dict[str, dict[str, Any]] = {}
    for hits in results_by_query.values():
        for hit in hits:
            k = hit.get("key", "")
            if k not in merged:
                merged[k] = dict(hit)
                merged[k]["match_count"] = 1
            else:
                merged[k]["score"] = merged[k]["score"] + hit.get("score", 0)
                merged[k]["match_count"] += 1

    ranked = sorted(merged.values(), key=lambda x: (-x.get("match_count", 1), -x.get("score", 0)))
    _bg.track_usage("memory_multi_search", _estimate_tokens(ranked))

    return {
        "queries": queries,
        "total_unique_results": len(ranked),
        "results": ranked,
    }


def memory_watch_vault(config: SynapseConfig, enable: bool = True) -> dict[str, Any]:
    """
    Watch the vault directory for external changes (e.g. Obsidian edits) and auto-rebuild
    the index when .md files are modified. enable=False stops the watcher.
    """
    _SENTINEL = "_vault_watcher"

    if not enable:
        flag = _bg.get_stop_flag(_SENTINEL)
        flag.set()
        return {"status": "stopped"}

    flag = _bg.get_stop_flag(_SENTINEL)
    if not flag.is_set():
        # Check if already running
        with _bg._lock:
            if _bg._jobs.get(_SENTINEL, {}).get("status") == "running":
                return {"status": "already_running"}

    flag.clear()

    vault = config.vault_path
    _ensure_vault_path(vault)

    def _watch() -> dict[str, Any]:
        import time as _time
        mtimes: dict[str, float] = {}

        def _snapshot() -> None:
            for p in vault.rglob("*.md"):
                if not p.name.startswith("_"):
                    try:
                        mtimes[str(p)] = p.stat().st_mtime
                    except OSError:
                        pass
        _snapshot()

        changes_detected = 0
        while not flag.is_set():
            _time.sleep(5)
            for p in list(vault.rglob("*.md")):
                if p.name.startswith("_") or flag.is_set():
                    continue
                try:
                    mtime = p.stat().st_mtime
                except OSError:
                    continue
                key = str(p)
                if mtimes.get(key) != mtime:
                    mtimes[key] = mtime
                    changes_detected += 1
                    _bg.schedule_rebuild(_rebuild_index_sync, config, delay_seconds=10)

        return {"changes_detected": changes_detected}

    _bg.start_job(_SENTINEL, _watch)
    return {"status": "started", "vault": str(vault), "note": "Polling every 5s. Call with enable=False to stop."}


def memory_ask(config: SynapseConfig, question: str, top_k: int = 5) -> dict[str, Any]:
    """
    Natural language Q&A over the vault. Retrieves relevant memories + chats, then asks
    Gemini to answer grounded in that context. Returns the answer with citations.
    Requires gemini_api_key in config.
    """
    if not config.gemini_api_key:
        return {"error": "gemini_api_key not set in config — required for memory_ask."}

    # Retrieve context: vault search + deep search if thin
    vault_hits = memory_search(config, question)[:top_k]
    deep_hits: list[dict[str, Any]] = []
    if len(vault_hits) < 2:
        try:
            deep_hits = _deep_search(config, question, depth=1, top_k=3)
        except Exception:
            pass

    # Build context block
    context_parts: list[str] = []
    citations: list[str] = []
    for hit in vault_hits:
        key = hit.get("key", "")
        snippet = hit.get("snippet") or hit.get("content", "")[:400]
        context_parts.append(f"[{key}]\n{snippet}")
        citations.append(key)
    for hit in deep_hits:
        key = hit.get("key") or hit.get("chat_id", "")
        snippet = hit.get("summary") or hit.get("snippet", "")[:400]
        context_parts.append(f"[{key}]\n{snippet}")
        citations.append(key)

    if not context_parts:
        return {"error": "No relevant memories found to answer this question. Try memory_search first."}

    context_text = "\n\n---\n\n".join(context_parts)
    prompt = (
        f"You are answering a question using only the memory records below. "
        f"Cite which record(s) support each claim using [key] notation.\n\n"
        f"MEMORIES:\n{context_text}\n\n"
        f"QUESTION: {question}\n\nANSWER:"
    )

    try:
        from google import genai as _genai
        client = _genai.Client(api_key=config.gemini_api_key)
        resp = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt,
            config={"temperature": 0.2},
        )
        answer = resp.text.strip()
    except Exception as e:
        return {"error": f"Gemini call failed: {e}"}

    tokens = _estimate_tokens({"answer": answer, "context": context_text})
    _bg.track_usage("memory_ask", tokens)

    return {
        "question": question,
        "answer": answer,
        "citations": citations,
        "context_sources": len(context_parts),
        "_tokens": tokens,
    }


def memory_health(config: SynapseConfig, auto_fix: bool = False) -> dict[str, Any]:
    """Vault health dashboard: file counts, token estimate, index age, graph stats, pending patches, issues. auto_fix=True runs deduplicate + relink + organize automatically."""
    import json as _json
    from datetime import datetime, timezone

    vault = config.vault_path
    _ensure_vault_path(vault)

    folder_counts: dict[str, int] = {}
    total_tokens = 0
    missing_frontmatter: list[str] = []

    for path in sorted(vault.rglob("*.md")):
        if path.name.startswith("_"):
            continue
        parts = path.relative_to(vault).parts
        folder = parts[0] if parts else "root"
        folder_counts[folder] = folder_counts.get(folder, 0) + 1
        try:
            text = read_text(config, path)
            total_tokens += len(text) // 4
            fm, _ = parse_memory_text(text)
            if not fm.get("type") or not fm.get("weight"):
                missing_frontmatter.append(path_to_key(vault, path))
        except Exception:
            missing_frontmatter.append(str(path.relative_to(vault)))

    total_files = sum(folder_counts.values())
    pending = list_pending(config)

    index_db = vault / "_index.db"
    index_age_hours = None
    if index_db.exists():
        age_secs = datetime.now(timezone.utc).timestamp() - index_db.stat().st_mtime
        index_age_hours = round(age_secs / 3600, 1)

    graph_stats: dict[str, Any] | None = None
    graph_file = vault / "metadata" / "topic_graph.json"
    if graph_file.exists():
        try:
            g = _json.loads(graph_file.read_text(encoding="utf-8"))
            graph_stats = {"nodes": len(g.get("nodes", [])), "edges": len(g.get("edges", []))}
        except Exception:
            pass

    issues: list[str] = []
    health_score = 100
    if missing_frontmatter:
        health_score -= min(20, len(missing_frontmatter) * 2)
        issues.append(f"{len(missing_frontmatter)} files with incomplete frontmatter")
    if len(pending) > 10:
        health_score -= 10
        issues.append(f"{len(pending)} patches pending review — run memory_diff")
    if index_age_hours is not None and index_age_hours > 24:
        health_score -= 10
        issues.append(f"Index is {index_age_hours}h old — run memory_rebuild_index")
    if not graph_stats:
        health_score -= 5
        issues.append("No topic graph built yet — run memory_build_graph")

    job_status = _bg.get_status()

    result: dict[str, Any] = {
        "health_score": max(0, health_score),
        "total_files": total_files,
        "total_tokens_est": total_tokens,
        "files_per_folder": folder_counts,
        "pending_patches": len(pending),
        "index_age_hours": index_age_hours,
        "graph": graph_stats,
        "missing_frontmatter": missing_frontmatter[:20],
        "issues": issues,
        "background_jobs": None if "message" in job_status else job_status,
    }

    if auto_fix and issues:
        fix_results: dict[str, Any] = {}
        try:
            fix_results["deduplicate"] = _dedup(config, auto_clean=True)
        except Exception as e:
            fix_results["deduplicate"] = {"error": str(e)}
        try:
            fix_results["relink"] = relink_all(config)
        except Exception as e:
            fix_results["relink"] = {"error": str(e)}
        try:
            fix_results["organize"] = organize_vault(config)
        except Exception as e:
            fix_results["organize"] = {"error": str(e)}
        _bg.schedule_rebuild(_rebuild_index_sync, config)
        result["auto_fix"] = fix_results

    return result


def memory_export_snapshot(config: SynapseConfig, output_path: str = "") -> dict[str, Any]:
    """Zip the vault to a timestamped snapshot file. Skips SQLite index files (rebuildable)."""
    import zipfile
    from datetime import datetime

    vault = config.vault_path
    _ensure_vault_path(vault)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = Path(output_path).expanduser().resolve() if output_path else config.root_path / f"synapse_snapshot_{ts}.zip"
    out.parent.mkdir(parents=True, exist_ok=True)

    skipped: list[str] = []
    file_count = 0
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(vault.rglob("*")):
            if not path.is_file():
                continue
            rel = path.relative_to(vault)
            if path.suffix in (".db", ".db-shm", ".db-wal"):
                skipped.append(str(rel))
                continue
            try:
                zf.write(path, arcname=f"vault/{rel}")
                file_count += 1
            except Exception as e:
                skipped.append(f"{rel} ({e})")

    return {
        "status": "exported",
        "snapshot_path": str(out),
        "files_included": file_count,
        "files_skipped": skipped,
        "size_bytes": out.stat().st_size,
    }


def memory_session_stats() -> dict[str, Any]:
    """Token budget and tool usage for the current session. Resets each server restart."""
    return _bg.get_session_stats()


def memory_session_save(config: SynapseConfig) -> dict[str, Any]:
    """
    Generate a prefilled memory_save_chat template based on this session's vault activity.
    Fill in title / summary / key_facts / decisions, then call memory_save_chat with the result.
    Note: MCP has no on-session-end hook, so this must be called explicitly before closing.
    """
    stats = _bg.get_session_stats()
    if not stats["keys_read"] and not stats["keys_written"]:
        return {
            "error": "No vault activity recorded yet this session. Use memory tools first, then call memory_session_save.",
        }

    touched = stats["keys_read"] + stats["keys_written"]
    folders = list(dict.fromkeys(k.split(".")[0] for k in touched if "." in k))

    return {
        "note": "Fill in title/summary/key_facts/decisions, then call memory_save_chat with this payload.",
        "template": {
            "title": "",
            "summary": "",
            "key_facts": [],
            "decisions": [],
            "tags": folders,
            "keywords": " ".join(folders),
            "categories": folders[:2] if folders else [],
        },
        "session_context": {
            "keys_read": stats["keys_read"],
            "keys_written": stats["keys_written"],
            "tokens_used": stats["tokens_used"],
            "elapsed_seconds": stats["session_elapsed_seconds"],
            "top_tools": stats["top_tools"],
        },
    }


def _ensure_vault_path(vault: Path) -> None:
    if not vault.exists():
        raise ValueError(f"Vault path does not exist: {vault}")
    if not vault.is_dir():
        raise ValueError(f"Vault path is not a directory: {vault}")


def _tree_node(path: Path, vault: Path) -> dict[str, Any]:
    children = []
    for child in sorted(path.iterdir(), key=lambda item: (not item.is_dir(), item.name.lower())):
        if child.name.startswith("."):
            continue
        if child.is_dir():
            children.append(_tree_node(child, vault))
        elif child.is_file():
            children.append(
                {
                    "type": "file",
                    "name": child.name,
                    "path": str(child.relative_to(vault)).replace("\\", "/"),
                    "key": path_to_key(vault, child) if child.suffix.lower() == ".md" else None,
                }
            )

    return {
        "type": "folder",
        "name": path.name,
        "path": "." if path == vault else str(path.relative_to(vault)).replace("\\", "/"),
        "children": children,
    }


# ── File ingestion ────────────────────────────────────────────────────────────

_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
_SUPPORTED_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".txt", ".md", ".html", ".htm",
    ".csv", ".tsv", ".xlsx", ".xls",
} | _IMAGE_EXTENSIONS

_IMAGE_MIME = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp",
}


def memory_read_file(config: SynapseConfig, file_path: str) -> dict[str, Any]:
    """
    Convert a file to markdown and return the text — no vault write, no patch.
    Use when Claude needs to read/analyze a document without saving it.
    Supports PDF, DOCX, XLSX, CSV, TSV, HTML, TXT, MD. Images are excluded.
    """
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        return {"error": f"File not found: {path}"}
    suffix = path.suffix.lower()
    if suffix in _IMAGE_EXTENSIONS:
        return {"error": "Images are not supported by memory_read_file. Use memory_preview_image instead."}
    if suffix not in _SUPPORTED_EXTENSIONS:
        return {"error": f"Unsupported file type: {suffix}", "supported": sorted(_SUPPORTED_EXTENSIONS - _IMAGE_EXTENSIONS)}
    try:
        markdown = _convert_to_markdown(path, config)
    except Exception as e:
        return {"error": f"Conversion failed: {e}"}
    tokens = _estimate_tokens({"markdown": markdown})
    _bg.track_usage("memory_read_file", tokens)
    return {
        "filename": path.name,
        "file_type": suffix.upper().lstrip("."),
        "characters": len(markdown),
        "markdown": markdown,
        "_tokens": tokens,
    }


def _convert_to_markdown(file_path: Path, config: SynapseConfig) -> str:
    """Convert any supported file type to markdown text. Images use Gemini vision."""
    suffix = file_path.suffix.lower()

    if suffix in (".md", ".txt"):
        return file_path.read_text(encoding="utf-8", errors="replace")

    if suffix == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(file_path))
            pages = []
            for i, page in enumerate(reader.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(f"## Page {i}\n\n{text.strip()}")
            return "\n\n".join(pages) if pages else "(PDF contained no extractable text)"
        except ImportError:
            raise ValueError("pypdf not installed — run: pip install pypdf")

    if suffix in (".docx", ".doc"):
        try:
            from docx import Document  # type: ignore
            doc = Document(str(file_path))
            lines = []
            for para in doc.paragraphs:
                style = para.style.name if para.style else ""
                text = para.text.strip()
                if not text:
                    lines.append("")
                    continue
                if style.startswith("Heading 1"):
                    lines.append(f"# {text}")
                elif style.startswith("Heading 2"):
                    lines.append(f"## {text}")
                elif style.startswith("Heading 3"):
                    lines.append(f"### {text}")
                else:
                    lines.append(text)
            return "\n\n".join(l for l in lines if l is not None)
        except ImportError:
            raise ValueError("python-docx not installed — run: pip install python-docx")

    if suffix in (".html", ".htm"):
        try:
            from markdownify import markdownify  # type: ignore
            html = file_path.read_text(encoding="utf-8", errors="replace")
            return markdownify(html, heading_style="ATX").strip()
        except ImportError:
            import re
            html = file_path.read_text(encoding="utf-8", errors="replace")
            return re.sub(r"<[^>]+>", "", html).strip()

    if suffix in (".csv", ".tsv"):
        import csv
        sep = "\t" if suffix == ".tsv" else ","
        rows = []
        with file_path.open(encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f, delimiter=sep)
            for i, row in enumerate(reader):
                rows.append("| " + " | ".join(row) + " |")
                if i == 0:
                    rows.append("|" + "|".join("---" for _ in row) + "|")
        return "\n".join(rows)

    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                parts.append(f"## Sheet: {sheet}\n")
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    cells = [str(c) if c is not None else "" for c in row]
                    rows.append("| " + " | ".join(cells) + " |")
                    if i == 0:
                        rows.append("|" + "|".join("---" for _ in cells) + "|")
                parts.append("\n".join(rows))
            return "\n\n".join(parts)
        except ImportError:
            raise ValueError("openpyxl not installed — run: pip install openpyxl")

    if suffix in _IMAGE_EXTENSIONS:
        raise ValueError(
            "Images must be ingested via memory_preview_image — "
            "call that tool first so Claude can assess sensitivity."
        )

    raise ValueError(f"Unsupported file type: {suffix}")


def memory_ingest_file(
    config: SynapseConfig,
    file_path: str,
    key: str = "",
    title: str = "",
    tags: list[str] | None = None,
) -> dict[str, Any]:
    """
    Convert a file (PDF, DOCX, XLSX, CSV, HTML, image, TXT) to markdown and
    store it in vault/files/ as a searchable memory. Returns the patch_id to apply.

    - file_path: absolute or relative path to the source file
    - key: vault key to use (default: files.<stem>)
    - title: display title (default: filename)
    - tags: optional list of topic tags stored as triggers
    """
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        return {"error": f"File not found: {path}"}
    if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
        return {
            "error": f"Unsupported file type: {path.suffix}",
            "supported": sorted(_SUPPORTED_EXTENSIONS),
        }

    if path.suffix.lower() in _IMAGE_EXTENSIONS:
        return {
            "status": "use_preview",
            "message": "Images require sensitivity review. Call memory_preview_image first.",
            "next_step": f"Call memory_preview_image('{file_path}')",
        }

    try:
        markdown = _convert_to_markdown(path, config)
    except Exception as e:
        return {"error": f"Conversion failed: {e}"}

    stem = path.stem.lower().replace(" ", "_").replace("-", "_")
    vault_key = key.strip() or f"files.{stem}"
    display_title = title.strip() or path.name
    trigger_list = list(tags or []) + [stem, path.suffix.lstrip(".")]

    # Build the memory body with file metadata header
    body = (
        f"# {display_title}\n\n"
        f"**Source:** `{path.name}`  \n"
        f"**Type:** {path.suffix.upper().lstrip('.')}  \n"
        f"**Ingested:** {__import__('datetime').date.today().isoformat()}\n\n"
        f"---\n\n"
        f"{markdown}"
    )

    from .memory_file import new_frontmatter as _new_fm
    fm = _new_fm(
        key=vault_key,
        memory_type="note",
        scope="global",
        weight=0.6,
        confidence="proposed",
        triggers=trigger_list,
    )

    try:
        result = propose_update(config, {
            "key": vault_key,
            "content": body,
            "merge": "replace",
            "frontmatter": fm,
            "reason": f"Ingested from file: {path.name}",
        })
    except Exception as e:
        return {"error": f"Failed to propose vault update: {e}"}

    _bg.track_usage("memory_ingest_file", _estimate_tokens({"body": body}), key=vault_key, write=True)

    return {
        "status": "proposed",
        "key": vault_key,
        "source_file": str(path),
        "file_type": path.suffix.upper().lstrip("."),
        "characters": len(markdown),
        "patch_id": result.get("patch_id"),
        "next_step": f"Call memory_apply_update('{result.get('patch_id')}') to save to vault.",
    }


def memory_list_files(config: SynapseConfig) -> dict[str, Any]:
    """
    List all files ingested into vault/files/ — the dedicated folder for file-converted memories.
    Returns key, title, source file name, type, and ingestion date for each.
    """
    vault = config.vault_path
    files_dir = vault / "files"
    if not files_dir.exists():
        return {"files": [], "count": 0, "note": "No files ingested yet. Use memory_ingest_file to add files."}

    entries = []
    for path in sorted(files_dir.glob("*.md")):
        if path.name.startswith("_"):
            continue
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
            fm, body = parse_memory_text(text)
            # Extract source file line from body
            source = ""
            for line in body.splitlines():
                if line.startswith("**Source:**"):
                    source = line.replace("**Source:**", "").strip().strip("`")
                    break
            entries.append({
                "key": fm.get("key", path_to_key(vault, path)),
                "title": body.splitlines()[0].lstrip("# ").strip() if body.strip() else path.stem,
                "source_file": source,
                "last_updated": fm.get("last_updated", ""),
                "triggers": fm.get("triggers", []),
            })
        except Exception:
            continue

    return {
        "files": entries,
        "count": len(entries),
        "folder": "vault/files/",
    }


# ── Content-first ingestion (file attached to conversation) ──────────────────

def _make_file_body(filename: str, content: str, file_type: str, source_label: str = "conversation attachment", title: str = "") -> str:
    import datetime as _dt
    heading = title.strip() or Path(filename).stem
    return (
        f"# {heading}\n\n"
        f"**Source:** `{filename}`  \n"
        f"**Type:** {file_type}  \n"
        f"**Origin:** {source_label}  \n"
        f"**Ingested:** {_dt.date.today().isoformat()}\n\n"
        f"---\n\n"
        f"{content.strip()}"
    )


def memory_ingest_file_content(
    config: SynapseConfig,
    filename: str,
    content: str,
    key: str = "",
    title: str = "",
    tags: list[str] | None = None,
) -> dict[str, Any]:
    """
    Ingest text content Claude received from a file attachment in the conversation.
    Use when the user drops a file into the chat and Claude already has the text —
    no filesystem path needed.

    - filename: original filename (e.g. 'report.pdf', 'notes.docx') — used for metadata
    - content:  the text content as Claude read it from the attachment
    - key:      vault key (default: files.<stem>)
    - title:    display title (default: filename stem)
    - tags:     optional topic tags
    """
    if not filename.strip():
        return {"error": "filename is required"}
    if not content.strip():
        return {"error": "content is empty — nothing to save"}

    path = Path(filename)
    stem = path.stem.lower().replace(" ", "_").replace("-", "_")
    file_type = path.suffix.upper().lstrip(".") or "FILE"
    vault_key = key.strip() or f"files.{stem}"
    trigger_list = list(tags or []) + [stem, path.suffix.lstrip(".").lower()]

    body = _make_file_body(filename, content, file_type, title=title)

    from .memory_file import new_frontmatter as _new_fm
    fm = _new_fm(key=vault_key, memory_type="note", scope="global",
                 weight=0.6, confidence="proposed", triggers=trigger_list)

    try:
        result = propose_update(config, {
            "key": vault_key,
            "content": body,
            "merge": "replace",
            "frontmatter": fm,
            "reason": f"File attached in conversation: {filename}",
        })
    except Exception as e:
        return {"error": f"Failed to propose vault update: {e}"}

    _bg.track_usage("memory_ingest_file_content", _estimate_tokens({"body": body}), key=vault_key, write=True)
    return {
        "status": "proposed",
        "key": vault_key,
        "filename": filename,
        "file_type": file_type,
        "characters": len(content),
        "patch_id": result.get("patch_id"),
        "next_step": f"Call memory_apply_update('{result.get('patch_id')}') to save to vault.",
    }


def memory_ingest_image_content(
    config: SynapseConfig,
    filename: str,
    description: str,
    key: str = "",
    title: str = "",
    tags: list[str] | None = None,
    sensitive: bool = False,
) -> dict[str, Any]:
    """
    Save an image Claude saw in the conversation using Claude's own description.
    Use when a user drops an image into the chat — Claude describes it here, no path needed.

    - filename:    original filename or descriptive name (e.g. 'screenshot.png')
    - description: Claude's detailed markdown description of the image
    - key:         vault key (default: files.<stem>)
    - title:       display title (default: filename stem)
    - tags:        optional topic tags
    - sensitive:   True if the image contained sensitive info (noted in metadata)
    """
    if not filename.strip():
        return {"error": "filename is required"}
    if not description.strip():
        return {"error": "description is empty — describe the image content first"}

    path = Path(filename)
    stem = path.stem.lower().replace(" ", "_").replace("-", "_")
    file_type = path.suffix.upper().lstrip(".") or "IMAGE"
    vault_key = key.strip() or f"files.{stem}"
    trigger_list = list(tags or []) + [stem, path.suffix.lstrip(".").lower()]

    source_label = "conversation attachment — described by Claude" + (" (sensitive)" if sensitive else "")
    body = _make_file_body(filename, description, file_type, source_label=source_label, title=title)

    from .memory_file import new_frontmatter as _new_fm
    fm = _new_fm(key=vault_key, memory_type="note", scope="global",
                 weight=0.7 if sensitive else 0.6,
                 confidence="confirmed", triggers=trigger_list)

    try:
        result = propose_update(config, {
            "key": vault_key,
            "content": body,
            "merge": "replace",
            "frontmatter": fm,
            "reason": f"Image attached in conversation: {filename}",
        })
    except Exception as e:
        return {"error": f"Failed to propose vault update: {e}"}

    _bg.track_usage("memory_ingest_image_content", _estimate_tokens({"body": body}), key=vault_key, write=True)
    return {
        "status": "proposed",
        "key": vault_key,
        "filename": filename,
        "sensitive": sensitive,
        "patch_id": result.get("patch_id"),
        "next_step": f"Call memory_apply_update('{result.get('patch_id')}') to save to vault.",
    }


# ── Image ingestion (Claude-routed) ──────────────────────────────────────────

def _image_preview_bytes(file_path: Path) -> tuple[bytes, str]:
    """Return (raw_bytes, format_string) for a supported image."""
    suffix = file_path.suffix.lower()
    fmt_map = {".png": "png", ".jpg": "jpeg", ".jpeg": "jpeg",
               ".gif": "gif", ".webp": "webp", ".bmp": "bmp"}
    return file_path.read_bytes(), fmt_map.get(suffix, "png")


def memory_preview_image(file_path: str) -> tuple[Any, dict[str, Any]]:
    """
    Load an image file and return it so Claude can see it directly in the conversation.
    After viewing, Claude should decide:
      - Not worth saving (blank, duplicate, no content) → do nothing
      - Contains sensitive info (passwords, medical, private messages, IDs, credentials)
        → call memory_ingest_image_save(file_path, markdown=<your own description>)
      - Safe to send to Gemini → call memory_ingest_image_gemini(file_path)
    Returns (Image, metadata_dict).
    """
    from mcp.server.fastmcp.utilities.types import Image as _MCPImage
    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        return {"error": f"File not found: {path}"}
    if path.suffix.lower() not in _IMAGE_EXTENSIONS:
        return {"error": f"Not an image file: {path.suffix}. Supported: {sorted(_IMAGE_EXTENSIONS)}"}

    raw, fmt = _image_preview_bytes(path)
    size_kb = round(len(raw) / 1024, 1)

    metadata = {
        "file": path.name,
        "size_kb": size_kb,
        "format": fmt.upper(),
        "instructions": (
            "You are viewing this image. Decide:\n"
            "1. NOT WORTH SAVING (blank, generic icon, no useful content) → do nothing, tell the user.\n"
            "2. SENSITIVE (passwords, credentials, medical records, private chats, ID documents, "
            "personal financial data) → call memory_ingest_image_save with your own markdown "
            "description (keeps data off Google).\n"
            "3. SAFE → call memory_ingest_image_gemini to let Gemini extract the markdown."
        ),
    }
    return _MCPImage(data=raw, format=fmt), metadata


def memory_ingest_image_save(
    config: SynapseConfig,
    file_path: str,
    markdown: str,
    key: str = "",
    title: str = "",
    tags: list[str] | None = None,
) -> dict[str, Any]:
    """
    Save an image to vault/files/ using Claude's own markdown description.
    Use this for sensitive images — no data is sent to external APIs.

    - file_path: path to the source image (used for metadata only)
    - markdown: Claude's description/extraction of the image
    - key: vault key (default: files.<stem>)
    - title: display title (default: filename)
    - tags: topic tags
    """
    path = Path(file_path).expanduser().resolve()
    stem = path.stem.lower().replace(" ", "_").replace("-", "_")
    vault_key = key.strip() or f"files.{stem}"
    display_title = title.strip() or path.name
    trigger_list = list(tags or []) + [stem, path.suffix.lstrip(".")]

    import datetime as _dt
    body = (
        f"# {display_title}\n\n"
        f"**Source:** `{path.name}`  \n"
        f"**Type:** {path.suffix.upper().lstrip('.')} (sensitive — extracted by Claude)  \n"
        f"**Ingested:** {_dt.date.today().isoformat()}\n\n"
        f"---\n\n"
        f"{markdown.strip()}"
    )

    from .memory_file import new_frontmatter as _new_fm
    fm = _new_fm(key=vault_key, memory_type="note", scope="global",
                 weight=0.7, confidence="confirmed", triggers=trigger_list)

    try:
        result = propose_update(config, {
            "key": vault_key,
            "content": body,
            "merge": "replace",
            "frontmatter": fm,
            "reason": f"Sensitive image ingested by Claude: {path.name}",
        })
    except Exception as e:
        return {"error": f"Failed to propose vault update: {e}"}

    _bg.track_usage("memory_ingest_image_save", _estimate_tokens({"body": body}), key=vault_key, write=True)
    return {
        "status": "proposed",
        "key": vault_key,
        "source": "claude",
        "patch_id": result.get("patch_id"),
        "next_step": f"Call memory_apply_update('{result.get('patch_id')}') to save.",
    }


def memory_ingest_image_gemini(
    config: SynapseConfig,
    file_path: str,
    key: str = "",
    title: str = "",
    tags: list[str] | None = None,
) -> dict[str, Any]:
    """
    Extract markdown from an image using Gemini vision and store in vault/files/.
    Use this for non-sensitive images. Requires gemini_api_key in config.

    - file_path: path to the source image
    - key: vault key (default: files.<stem>)
    - title: display title (default: filename)
    - tags: topic tags
    """
    if not config.gemini_api_key:
        return {"error": "gemini_api_key not set in config. For sensitive images use memory_ingest_image_save instead."}

    path = Path(file_path).expanduser().resolve()
    if not path.exists():
        return {"error": f"File not found: {path}"}
    if path.suffix.lower() not in _IMAGE_EXTENSIONS:
        return {"error": f"Not an image file: {path.suffix}"}

    import base64
    from google import genai as _genai
    client = _genai.Client(api_key=config.gemini_api_key)
    raw = path.read_bytes()
    b64 = base64.b64encode(raw).decode()
    mime = _IMAGE_MIME.get(path.suffix.lower(), "image/png")

    try:
        resp = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=[
                {"inline_data": {"mime_type": mime, "data": b64}},
                "Extract all text and describe the contents of this image in well-structured markdown. "
                "Use headings for sections, bullet points for lists, and preserve any tables.",
            ],
            config={"temperature": 0.1},
        )
        markdown = resp.text.strip()
    except Exception as e:
        return {"error": f"Gemini extraction failed: {e}"}

    stem = path.stem.lower().replace(" ", "_").replace("-", "_")
    vault_key = key.strip() or f"files.{stem}"
    display_title = title.strip() or path.name
    trigger_list = list(tags or []) + [stem, path.suffix.lstrip(".")]

    import datetime as _dt
    body = (
        f"# {display_title}\n\n"
        f"**Source:** `{path.name}`  \n"
        f"**Type:** {path.suffix.upper().lstrip('.')} (extracted by Gemini)  \n"
        f"**Ingested:** {_dt.date.today().isoformat()}\n\n"
        f"---\n\n"
        f"{markdown}"
    )

    from .memory_file import new_frontmatter as _new_fm
    fm = _new_fm(key=vault_key, memory_type="note", scope="global",
                 weight=0.6, confidence="proposed", triggers=trigger_list)

    try:
        result = propose_update(config, {
            "key": vault_key,
            "content": body,
            "merge": "replace",
            "frontmatter": fm,
            "reason": f"Image ingested via Gemini: {path.name}",
        })
    except Exception as e:
        return {"error": f"Failed to propose vault update: {e}"}

    _bg.track_usage("memory_ingest_image_gemini", _estimate_tokens({"body": body}), key=vault_key, write=True)
    return {
        "status": "proposed",
        "key": vault_key,
        "source": "gemini",
        "characters": len(markdown),
        "patch_id": result.get("patch_id"),
        "next_step": f"Call memory_apply_update('{result.get('patch_id')}') to save.",
    }
